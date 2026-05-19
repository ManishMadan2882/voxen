import io
from typing import Iterable

from openpyxl import load_workbook

from rag.document_model import DocType
from rag.ingestors.base import FileIngestor, ParsedSegment


class XlsxIngestor(FileIngestor):
    extensions = (".xlsx", ".xls")
    doc_type = DocType.xlsx

    def parse(self, content: bytes, filename: str) -> Iterable[ParsedSegment]:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        # one logical "page" per sheet so retrieval can cite a sheet name
        for sheet_idx, ws in enumerate(wb.worksheets):
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v).strip() for v in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            text = "\n".join(rows).strip()
            if not text:
                continue
            source = f"{filename}#{ws.title}" if ws.title else filename
            yield ParsedSegment(source=source, page=sheet_idx + 1, text=text)
